import requests
from bs4 import BeautifulSoup
import openpyxl
import os
from openpyxl.drawing.image import Image


#Загрузка изображений для последующей вставки
def download_images(img_url, art):
    # Получение пути к текущей директории, где находится исполняемый файл
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # Создание относительного пути к папке для изображений
    img_dir = os.path.join(current_dir, "img")
    # Создание директории, если её нет
    os.makedirs(img_dir, exist_ok=True)
    # Сохранение изображения
    img_path = os.path.join(img_dir, f"{art}.jpg")
    with open(img_path, "wb") as f:
        f.write(requests.get(img_url).content)
    return 1

# Загрузка страницы и извлечение данных
def extract_data(url):
    response = requests.get(url)
    # print(response.content)
    soup = BeautifulSoup(response.content, 'html.parser')
    # Извлечение заголовка товара по id
    title = soup.find(id='title').text.strip()
    # Извлечение артикула товара по class
    art = soup.find(class_='line ttl').text.strip()
    arts = art.removeprefix("Артикул: ")
    # Извлечение изображения товара через название класса материнского элемента div и тега img
    image_div = soup.find('div', class_='sl-item')
    image_div_inner = image_div.find('div', class_='pic')
    image_src = image_div_inner.find('img')['src']
    image_link = 'https://www.officeton.by' + image_src
    download_images(image_link, arts)
    # Извлечение цены товара по id
    try:
        price = soup.find('span', class_='price-item__val price-item__val_new').text.strip()
    except AttributeError:
        price = 'Нет в наличии'
    return title, image_link, arts, price

#Вставка изображения в таблицу
def image_insert(sheet, articul, index):
    # Получение пути к текущей директории, где находится исполняемый файл
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # Создание относительного пути к папке с изображениями
    img_dir = os.path.join(current_dir, "img")
    image_path = os.path.join(img_dir, f"{articul}.jpg")
    img = Image(image_path)
    img.height = 150  # Высота в пикселях
    img.width = 150  # Ширина в пикселях
    sheet2.row_dimensions[index+2].height = 130
    cell_coordinates = f'B{index+2}'
    sheet.add_image(img, cell_coordinates)
    return 1

# Загрузка Excel-таблицы
workbook = openpyxl.load_workbook(input("введите название excel-файла с ссылками:")+".xlsx")
sheet = workbook['Ссылки']
if "позиции" not in workbook.sheetnames:
    sheet2 = workbook.create_sheet("позиции")  # посмотреть,отутствие листа перед запуском

else:
    sheet2 = workbook['позиции']
    workbook.remove(sheet2)
    workbook.save('Links.xlsx')
    sheet2 = workbook.create_sheet("позиции")  # посмотреть,отутствие листа перед запуском



i = 1
# Перебор ссылок в каждой строке таблицы
sheet2.append({'A': '', 'C': '', 'D': '', 'E': ''})
sheet2.append({'A': 'Название', 'C': 'Ссылка', 'D': 'Артикул', 'E': 'Цена'})

print('\n')
for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
    url = row[0]  # выбор ссылки на товар
    if url!=None:
        url=url.replace('\n','')
        title, image, art, price = extract_data(url)
        print('Ссылка', url)
        print('Название', title)
        print('Артикул', art)
        print('Цена', price, '\n')
        # Запись данных в следующие столбцы таблицы
        sheet2.append({'A': title, 'C': url, 'D': art, 'E': price})
        image_insert(sheet2, art, i)
        i += 1
        sheet2.column_dimensions['B'].width = 33
workbook.save('Links.xlsx')
input('Список позиций сформирован, нажмите Enter, чтобы завершить.')

